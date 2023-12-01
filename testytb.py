from tkinter import filedialog
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl.drawing.image import Image
import pyperclip
import time
from selenium import webdriver
import datetime
import pyautogui as pi
import os
import json
import cv2





class Scrappy:
    options = webdriver.ChromeOptions()
    options.add_experimental_option(
            'excludeSwitches', ['enable-logging']
        )
    options.add_argument('--lang=pr-BR')
    options.add_argument('--disable-notifications')
    driver = webdriver.Chrome(options=options)
      
    #criação dos arrays de dados
    data = []

    
    
    def iniciar(self):
        self.raspagemDeDadosCanadian()
        self.criarPlanilha()
    #estrutura de raspagem de dados de acordo com o site (login(se nescessario), paths das informações a ser consumidas)
    def raspagemDeDadosCanadian(self):
        self.link = 'https://monitoring.csisolar.com/maintain/plant'
        self.driver.get(self.link)
        self.driver.find_element(By.XPATH, '//*[@id="app"]/div[5]/div[3]/div/div[4]/button[1]').click()
        self.driver.find_element(By.XPATH, '//*[@id="app"]/div[5]/div[5]/div/section/div[2]/div[2]/div[3]/div/input').send_keys('operacao@energiaarion.com.br', Keys.TAB, 'Arion2023', Keys.ENTER)
        time.sleep(10)
        self.driver.find_element(By.XPATH, '//*[@id="app"]/div[5]/div[2]/section/div/section/div[4]/div[1]/div[1]/div[2]/a').click()
        print('\033[42m'+'\033[1m'+'\033[33m'+'Vamos começar a varredura CANADIAN☀️'+'\033[0;0m')
        value = 1
        time.sleep(5)
        for n in range(1):
            for i in range(16):
                lista_localizacao = self.driver.find_elements(By.XPATH , f'/html/body/div[1]/div[5]/div[2]/section/div/div[5]/div[2]/div[1]/div[1]/div[1]/div[2]/div[1]/div/table/tbody/tr[{value}]/td[2]/div[1]')
                self.data.append(lista_localizacao[0].get_attribute('innerHTML'))
                value += 1
        self.driver.quit()        
        print(self.data)
        print('\033[42m'+'\033[1m'+'\033[33m'+'varredura terminou☀️'+'\033[0;0m')
        time.sleep(10)
   
    #criação da planilha com as informações colidas no site
    def criarPlanilha(self):
        planilha = openpyxl.Workbook()
        geracao = planilha['Sheet']
        geracao.title = 'Usinas Criativa Energia'
        geracao['A1'] = 'Nome'
        geracao['B1'] = 'Endereço'
        geracao['C1'] = 'Geração Diaria'
        index = 2
        elemento = ','
        for localizacao, end, geracaoDiaria in zip(self.listaLocCa, self.listaEndCa, self.listaGeracaoCa):
            geracao.cell(column=1, row=index, value=localizacao)
            geracao.cell(column=2, row=index,  value=end)
            geracao.cell(column=3, row=index, value=fr'=VALOR({geracaoDiaria})')
            print(f'=VALOR({geracaoDiaria})')
            geracao.cell(column=4, row=index, value='https://static.semsportal.com/static/Images/icon-weather-new/101.png')
            print(f'=IMAGEM("D{index}";"CLIMA")')
            geracao.cell(column=5, row=index, value=f'=IMAGEM(D{index};"CLIMA")')
            geracao.cell(column=6, row=index, value=f'{self.dataStr}, {hora}')
            index += 1
        index = 18                    
        print('\033[32m'+'Planilha ----------- 33,33% '+'\033[0;0m')    
        
        planilha.save('Raspagem.xlsx')
        print('\033[32m'+'Planilha criada com sucesso☀️🚩------------------------------------------100%😜'+'\033[0;0m') 
        pi.alert("O PROCESSO DE ATUALIZAÇÃO TERMINOU ☀️🚩😜!! A PLANILHA ESTÁ NA MESMA PASTA DO AQUIVO DO CODIGO 😜🧧✅")

    
                

start = Scrappy()
start.iniciar()